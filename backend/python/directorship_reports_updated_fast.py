import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import datetime
from zipfile import ZipFile
import pytz
import subprocess
import shutil
from concurrent.futures import ProcessPoolExecutor, as_completed
import multiprocessing
import traceback

# ============================================================
#                HARD-CODED USER VARIABLES
# ============================================================

INPUT_CSV = "/content/drive/MyDrive/LIILMMRA/LLMTech/UAN Bulk Automation Script/Directorship Check Script/Output/output_results.csv"

PDF_OUTPUT_DIR = "/content/drive/MyDrive/LIILMMRA/LLMTech/UAN Bulk Automation Script/Directorship Check Script/PDF Reports/"
ZIP_OUTPUT_DIR = "/content/drive/MyDrive/LIILMMRA/LLMTech/UAN Bulk Automation Script/Directorship Check Script/Zip Files/"

# Column widths
COL_WIDTH_PAN = 15
COL_WIDTH_NAME = 16
COL_WIDTH_DIN = 20
COL_WIDTH_GST = 28
COL_WIDTH_STATUS = 8

# Colors
COLOR_TITLE = "BDD7EE"   # Light Cornflower Blue 2
COLOR_HEADER = "C6E0B4"  # Light Green 2

IST = pytz.timezone("Asia/Kolkata")


# ============================================================
# HELPERS
# ============================================================

def process_single_row(args):
    """
    Worker-safe function.
    Handles ONE row → Excel → PDF.
    """

    (
        row_dict,
        soffice_path,
        temp_excel,
        pdf_folder,
        index,
        total
    ) = args

    try:
        ref_no = sanitize_filename(str(row_dict["Ref No"]))
        candidate = sanitize_filename(row_dict["Candidate Name"])
        base_name = f"{ref_no}-{candidate}"

        excel_path = os.path.join(temp_excel, f"{base_name}.xlsx")
        pdf_path = os.path.join(pdf_folder, f"{base_name}.pdf")

        print(f"[{index}/{total}] 📄 Creating Excel → {base_name}")
        create_excel_report(row_dict, excel_path, candidate)

        print(f"[{index}/{total}] 🔄 Converting to PDF → {base_name}")
        convert_excel_to_pdf(soffice_path, excel_path, pdf_path)

        print(f"[{index}/{total}] ✅ Done → {base_name}")

        return pdf_path, None

    except Exception as e:
        error_msg = (
            f"\n❌ ERROR processing Ref No {row_dict.get('Ref No')} "
            f"({row_dict.get('Candidate Name')})\n"
            f"{traceback.format_exc()}"
        )
        return None, error_msg


def adjust_row_height_for_name_and_gst(ws, row_num, base_height=18):
    """
    Increase row height based on max lines in
    Name (B) and GST (D) columns.
    """
    max_lines = 1

    for col_letter in ["B", "D"]:
        value = ws[f"{col_letter}{row_num}"].value
        if not value:
            continue

        lines = str(value).replace(",", "\n").split("\n")
        max_lines = max(max_lines, len([l for l in lines if l.strip()]))

    ws.row_dimensions[row_num].height = base_height * max_lines

def adjust_gst_row_height(ws, row_num, col_letter="D", base_height=18):
    """
    Dynamically increases row height based on number of GST entries.
    Assumes GST values are separated by commas or newlines.
    """
    cell_value = ws[f"{col_letter}{row_num}"].value

    if not cell_value:
        return

    # Count GST lines (commas or newlines)
    lines = str(cell_value).replace(",", "\n").split("\n")
    line_count = max(1, len([l for l in lines if l.strip()]))

    ws.row_dimensions[row_num].height = base_height * line_count

def sanitize_filename(name):
    invalid = '<>:"/\\|?*'
    for c in invalid:
        name = name.replace(c, "")
    return name.strip()

def ensure_dir(path):
    if not os.path.exists(path):
        os.makedirs(path)

def get_timestamp():
    now = datetime.now(IST)
    return now.strftime("%d/%m/%Y %H:%M:%S")

def require_soffice():
    soffice_path = shutil.which("soffice")
    if soffice_path is None:
        raise RuntimeError(
            "\n❌ LibreOffice not installed.\n"
            "Install inside Codespaces:\n"
            "  sudo apt update\n"
            "  sudo apt install libreoffice --no-install-recommends\n"
        )
    print(f"✔ soffice detected: {soffice_path}")
    return soffice_path


# ============================================================
#       CREATE DIRECTORSHIP REPORT EXCEL (v6)
# ============================================================

def create_excel_report(row, excel_path, candidate_name):

    wb = Workbook()
    ws = wb.active

    # Pre-create rows to avoid corruption
    for _ in range(10):
        ws.append([])

    # Column widths
    ws.column_dimensions["A"].width = COL_WIDTH_PAN
    ws.column_dimensions["B"].width = COL_WIDTH_NAME
    ws.column_dimensions["C"].width = COL_WIDTH_DIN
    ws.column_dimensions["D"].width = COL_WIDTH_GST
    ws.column_dimensions["E"].width = COL_WIDTH_STATUS

    # Borders & alignment
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center")

    # ====================================================
    #  HEADER (Excel printed header)
    # ====================================================
    ws.oddHeader.left.text = candidate_name

    # ====================================================
    #  TITLE (Row 1)
    # ====================================================
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "MCA - Directorship Report"
    t.font = Font(name="Arial", size=12, bold=True)
    t.fill = PatternFill(start_color=COLOR_TITLE, end_color=COLOR_TITLE, fill_type="solid")
    t.alignment = center

    # Apply border to all merged title cells
    for col in range(1, 6):
        ws.cell(row=1, column=col).border = border

    # Row 2 → Blank
    ws["A2"] = ""

    # ====================================================
    #  TABLE HEADER (Row 3)
    # ====================================================
    headers = ["PAN", "Full Name", "DIN", "GST", "Status"]

    for col, text in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=text)
        c.font = Font(name="Arial", size=9, bold=True)
        c.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        c.border = border
        c.alignment = center


    # ====================================================
    #  TABLE ROW (Row 4)
    # ====================================================
    pan = row["PAN"]
    name = row["Candidate Name"]
    din = "No records found" if row["DIN"] == "N" else row["DIN"]
    gst = "No records found" if row["GST"] == "N" else row["GST"].replace(",", "\n")
    status = row["Status"]

    values = [pan, name, din, gst, status]

    # Write values + alignment
    for col, value in enumerate(values, 1):
        c = ws.cell(row=4, column=col, value=value)
        c.font = Font(name="Arial", size=9)
        c.border = border

        c.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=(col in (2, 4))
        )

    # Adjust GST row height ONCE
    adjust_row_height_for_name_and_gst(ws, row_num=4)


    # 🔒 FORCE CENTER ALIGNMENT FOR ENTIRE ROW (LibreOffice-safe)
    for col in range(1, 6):
        ws.cell(row=4, column=col).alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=(col in (2, 4))
        )


    # Row 5 → Blank
    ws["A5"] = ""

    # ====================================================
    #  URL ROW (Row 6)
    # ====================================================
    url = ("https://www.mca.gov.in/content/mca/global/en/mca/master-data/"
           "View-Companies-Directors-under-prosecution-V3.html")

    ws.merge_cells("A6:E6")
    u = ws["A6"]
    u.value = url
    u.font = Font(name="Arial", size=8)
    u.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )
    ws.row_dimensions[6].height = 45  # prevents cut-off

    # Row 7 → Blank
    ws["A7"] = ""

    # ====================================================
    #  FOOTER (Printed footer)
    # ====================================================
    timestamp = get_timestamp()
    ws.oddFooter.left.text = f"Generated: {timestamp}"
    ws.oddFooter.right.text = "&P"  # page number bottom-right

    # ====================================================
    #  PAGE SETUP (matches your working EPFO style)
    # ====================================================
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9  # A4

    ws.page_setup.fitToWidth = True
    ws.page_setup.fitToHeight = False
    ws.page_setup.fitToPage = True

    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3

    ws.print_area = "A1:E8"

    # Center entire printed content
    try:
        ws.print_options.horizontalCentered = True
    except:
        pass

    wb.save(excel_path)


# ============================================================
#     EXCEL → PDF (Scan & Rename Safe Method)
# ============================================================

import time
import glob

def convert_excel_to_pdf(soffice_path, excel_path, pdf_path):
    """
    Colab-safe LibreOffice PDF conversion.
    Avoids directory diff logic which fails in Colab.
    """

    output_dir = os.path.dirname(excel_path)
    base_name = os.path.splitext(os.path.basename(excel_path))[0]

    cmd = [
        soffice_path,
        "--headless",
        "--nologo",
        "--nofirststartwizard",
        "--convert-to", "pdf:calc_pdf_Export",
        excel_path,
        "--outdir", output_dir
    ]

    subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)

    # LibreOffice writes asynchronously in Colab
    time.sleep(2)

    # Search for generated PDF
    candidates = glob.glob(os.path.join(output_dir, f"{base_name}*.pdf"))

    if not candidates:
        raise RuntimeError(
            f"❌ LibreOffice did not generate PDF.\n"
            f"Excel: {excel_path}\n"
            f"Checked: {output_dir}"
        )

    generated_pdf = candidates[0]

    if generated_pdf != pdf_path:
        shutil.move(generated_pdf, pdf_path)



# ============================================================
#                       MAIN DRIVER
# ============================================================

def generate_reports():

    soffice_path = require_soffice()

    df = pd.read_csv(INPUT_CSV)
    today = datetime.now(IST).strftime("%d-%m-%Y")

    pdf_folder = os.path.join(PDF_OUTPUT_DIR, f"DC {today}")
    ensure_dir(pdf_folder)

    temp_excel = os.path.join(pdf_folder, "temp_excel")
    ensure_dir(temp_excel)

    ensure_dir(ZIP_OUTPUT_DIR)
    zip_path = os.path.join(ZIP_OUTPUT_DIR, f"DC_{today}.zip")

    total = len(df)

    print(f"\n🚀 Starting Directorship Report Generation")
    print(f"📊 Total Records: {total}\n")

    # ---- Worker count (safe for LibreOffice) ----
    max_workers = min(
        max(1, multiprocessing.cpu_count() - 1),
        6  # HARD CAP → LibreOffice stability
    )

    print(f"⚙ Using {max_workers} parallel workers\n")

    tasks = []
    for idx, (_, row) in enumerate(df.iterrows(), start=1):
        tasks.append((
            row.to_dict(),
            soffice_path,
            temp_excel,
            pdf_folder,
            idx,
            total
        ))

    pdf_files = []
    errors = []

    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(process_single_row, t) for t in tasks]

        for future in as_completed(futures):
            pdf, error = future.result()
            if pdf:
                pdf_files.append(pdf)
            if error:
                errors.append(error)
                print(error)

    # ---- ZIP CREATION ----
    print("\n🗜 Creating ZIP file...")

    with ZipFile(zip_path, "w") as zipf:
        for pdf in pdf_files:
            zipf.write(pdf, os.path.basename(pdf))

    shutil.rmtree(temp_excel, ignore_errors=True)

    print("\n🎉 REPORT GENERATION COMPLETE")
    print(f"📄 PDFs Generated: {len(pdf_files)} / {total}")
    print(f"📂 PDF Folder: {pdf_folder}")
    print(f"🗜 ZIP File:   {zip_path}")

    if errors:
        print(f"\n⚠ Completed with {len(errors)} errors (see logs above)")



# ============================================================
#                     ENTRY POINT
# ============================================================

if __name__ == "__main__":
    generate_reports()