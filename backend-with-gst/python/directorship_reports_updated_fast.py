import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import datetime
from zipfile import ZipFile
import pytz
import subprocess
import shutil
import argparse
import time
import glob

# ============================================================
# CONSTANTS
# ============================================================

COL_WIDTH_PAN = 18
COL_WIDTH_NAME = 22
COL_WIDTH_DIN = 24
COL_WIDTH_GST = 32
COL_WIDTH_STATUS = 10

COLOR_TITLE = "BDD7EE"
COLOR_HEADER = "C6E0B4"

IST = pytz.timezone("Asia/Kolkata")

# ============================================================
# HELPERS
# ============================================================

def sanitize_filename(name):
    invalid = '<>:"/\\|?*'
    for c in invalid:
        name = name.replace(c, "")
    return name.strip()

def ensure_dir(path):
    if not os.path.exists(path):
        os.makedirs(path)

def get_timestamp():
    now = datetime.now(IST)
    return now.strftime("%d/%m/%Y %H:%M:%S")

def require_soffice():
    possible_paths = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]
    for path in possible_paths:
        if os.path.exists(path):
            print(f"[OK] soffice detected at: {path}", flush=True)
            return path
    raise RuntimeError("LibreOffice not found")

# ============================================================
# SMART ROW HEIGHT (FIXED)
# ============================================================

def adjust_dynamic_row_height(ws, row_num, base_height=18):

    max_lines = 1

    approx_chars = {
        "B": 22,   # Name
        "C": 25,   # DIN
        "D": 35    # GST
    }

    for col_letter in ["B", "C", "D"]:
        cell = ws[f"{col_letter}{row_num}"]
        value = cell.value

        if not value:
            continue

        text = str(value).replace(",", "\n")
        lines = text.split("\n")

        estimated_lines = 0

        for line in lines:
            if not line.strip():
                continue
            max_per_line = approx_chars.get(col_letter, 25)
            estimated_lines += max(1, len(line) // max_per_line + 1)

        max_lines = max(max_lines, estimated_lines)

    ws.row_dimensions[row_num].height = base_height * max_lines

# ============================================================
# OLD UI DESIGN (ENHANCED)
# ============================================================

def create_excel_report(row, excel_path, candidate_name):

    wb = Workbook()
    ws = wb.active

    for _ in range(10):
        ws.append([])

    ws.column_dimensions["A"].width = COL_WIDTH_PAN
    ws.column_dimensions["B"].width = COL_WIDTH_NAME
    ws.column_dimensions["C"].width = COL_WIDTH_DIN
    ws.column_dimensions["D"].width = COL_WIDTH_GST
    ws.column_dimensions["E"].width = COL_WIDTH_STATUS

    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center")

    ws.oddHeader.left.text = candidate_name

    # TITLE
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "MCA - Directorship Report"
    t.font = Font(name="Arial", size=12, bold=True)
    t.fill = PatternFill(start_color=COLOR_TITLE, end_color=COLOR_TITLE, fill_type="solid")
    t.alignment = center

    for col in range(1, 6):
        ws.cell(row=1, column=col).border = border

    # TABLE HEADER
    headers = ["PAN", "Full Name", "DIN", "GST", "Status"]

    for col, text in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=text)
        c.font = Font(name="Arial", size=9, bold=True)
        c.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        c.border = border
        c.alignment = center

    # DATA
    pan = row["PAN"]
    name = row["Candidate Name"]

    din = "No records found" if row["DIN"] == "N" else row["DIN"]
    gst = "No records found" if row["GST"] == "N" else row["GST"].replace(",", "\n")
    status = row["Status"]

    values = [pan, name, din, gst, status]

    for col, value in enumerate(values, 1):
        c = ws.cell(row=4, column=col, value=value)
        c.font = Font(name="Arial", size=9)
        c.border = border
        c.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        # Status color
        if col == 5:
            c.font = Font(name="Arial", size=9, bold=True, color="000000")

    # Dynamic height
    adjust_dynamic_row_height(ws, 4)

   # URL (Clickable Hyperlink)

   
    url = ("https://www.mca.gov.in/content/mca/global/en/mca/master-data/"
        "View-Companies-Directors-under-prosecution-V3.html")

    ws.merge_cells("A6:E6")
    u = ws["A6"]

    u.value = url
    u.hyperlink = url
    u.style = "Hyperlink"

    u.font = Font(
        name="Arial",
        size=8,
        underline="single",
        color="0563C1"  # Standard hyperlink blue
    )

    u.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    ws.row_dimensions[6].height = 45


    # Footer
    ws.oddFooter.left.text = f"Generated: {get_timestamp()}"
    ws.oddFooter.right.text = "&P"

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToWidth = True
    ws.page_setup.fitToHeight = False
    ws.page_setup.fitToPage = True

    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75

    ws.print_options.horizontalCentered = True

    ws.print_area = "A1:E8"

    wb.save(excel_path)

# ============================================================
# SAFE PDF CONVERSION
# ============================================================

def convert_excel_to_pdf(soffice_path, excel_path, pdf_path):

    # Validate input file
    if not os.path.exists(excel_path):
        raise RuntimeError(f"Excel file does not exist: {excel_path}")

    if not os.access(excel_path, os.R_OK):
        raise RuntimeError(f"Excel file is not readable: {excel_path}")

    # Check file size (empty files might cause issues)
    if os.path.getsize(excel_path) == 0:
        raise RuntimeError(f"Excel file is empty: {excel_path}")

    output_dir = os.path.dirname(excel_path)
    base_name = os.path.splitext(os.path.basename(excel_path))[0]

    # Clean up any existing PDF files with the same base name
    existing_pdfs = glob.glob(os.path.join(output_dir, f"{base_name}*.pdf"))
    for pdf in existing_pdfs:
        try:
            os.remove(pdf)
        except:
            pass  # Ignore cleanup errors

    cmd = [
        soffice_path,
        "--headless",
        "--nologo",
        "--nofirststartwizard",
        "--convert-to", "pdf:calc_pdf_Export",
        excel_path,
        "--outdir", output_dir
    ]

    max_retries = 3
    for attempt in range(max_retries):
        try:
            result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=30)

            if result.returncode != 0:
                error_msg = f"LibreOffice conversion failed with code {result.returncode}"
                if result.stderr:
                    error_msg += f"\nStderr: {result.stderr}"
                if attempt < max_retries - 1:
                    print(f"Attempt {attempt + 1} failed, retrying... Error: {error_msg}", flush=True)
                    time.sleep(2)
                    continue
                else:
                    raise RuntimeError(f"{error_msg}\nCommand: {' '.join(cmd)}")

            # Wait for file to be written
            time.sleep(3)

            candidates = glob.glob(os.path.join(output_dir, f"{base_name}*.pdf"))

            if candidates:
                generated_pdf = candidates[0]
                if generated_pdf != pdf_path:
                    shutil.move(generated_pdf, pdf_path)
                return  # Success

            if attempt < max_retries - 1:
                print(f"Attempt {attempt + 1}: PDF not found, retrying...", flush=True)
                time.sleep(2)
            else:
                raise RuntimeError(f"PDF not generated for {excel_path} after {max_retries} attempts\nLibreOffice stdout: {result.stdout}")

        except subprocess.TimeoutExpired:
            if attempt < max_retries - 1:
                print(f"Attempt {attempt + 1} timed out, retrying...", flush=True)
                time.sleep(2)
            else:
                raise RuntimeError(f"LibreOffice conversion timed out for {excel_path} after {max_retries} attempts")

# ============================================================
# MAIN DRIVER
# ============================================================

def generate_reports(input_csv, output_zip):

    soffice_path = require_soffice()
    df = pd.read_csv(input_csv)

    job_dir = os.path.dirname(output_zip)

    reports_folder = os.path.join(job_dir, "reports")
    ensure_dir(reports_folder)

    total = len(df)

    print(f"\n[START] Generating Reports", flush=True)
    print(f"[INFO] Total Records: {total}\n", flush=True)

    pdf_files = []
    failed_conversions = []

    for index, row in df.iterrows():

        ref_no = sanitize_filename(str(row["Ref No"]))
        candidate = sanitize_filename(row["Candidate Name"])
        base_name = f"{ref_no}-{candidate}"

        excel_path = os.path.join(reports_folder, f"{base_name}.xlsx")
        pdf_path = os.path.join(reports_folder, f"{base_name}.pdf")

        print(f"[{index+1}/{total}] Creating Excel -> {base_name}", flush=True)
        create_excel_report(row, excel_path, candidate)

        print(f"[{index+1}/{total}] Converting to PDF -> {base_name}", flush=True)
        try:
            convert_excel_to_pdf(soffice_path, excel_path, pdf_path)
            pdf_files.append(pdf_path)
            print(f"[{index+1}/{total}] PDF Generated Successfully -> {base_name}", flush=True)
        except Exception as e:
            error_msg = f"Failed to convert {base_name}: {str(e)}"
            print(f"[ERROR] {error_msg}", flush=True)
            failed_conversions.append((base_name, str(e)))
            # Continue with next file instead of crashing

    if failed_conversions:
        print(f"\n[WARNING] {len(failed_conversions)} PDF conversions failed:", flush=True)
        for name, error in failed_conversions:
            print(f"  - {name}: {error}", flush=True)

    print("\n[ZIP] Creating ZIP file...", flush=True)

    with ZipFile(output_zip, "w") as zipf:
        for pdf in pdf_files:
            zipf.write(pdf, os.path.basename(pdf))

    print("\n[SUCCESS] REPORT GENERATION COMPLETE", flush=True)
    print(f"[PDF] Generated: {len(pdf_files)} / {total}", flush=True)
    if failed_conversions:
        print(f"[FAILED] PDF conversions: {len(failed_conversions)} / {total}", flush=True)
    print(f"[DIR] {reports_folder}", flush=True)
    print(f"[ZIP] {output_zip}", flush=True)

# ============================================================
# ENTRY
# ============================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("input_csv")
    parser.add_argument("output_zip")
    args = parser.parse_args()

    generate_reports(args.input_csv, args.output_zip)
